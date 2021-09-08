import openpyxl as pxl
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
import copy
import re
import os


class Excel:
    """
    将Excel内存转成内容——位置（eg:A1）对照的字典
    未作异常处理
    """
    def __init__(self, filePath, readOnly=False):
        self.filepath = filePath
        self.readWrite = readOnly
        self.keyValue = dict()
        self.cellLength = dict()
        self.ColumnLength = dict()
        self.workbook = Excel.openOrCreate(self, filePath)

    def openOrCreate(self, filePath: str):
        """若有则打开，若无则创建。

        Args:
            path (str):文件路径
        """
        if os.path.exists(filePath):
            pass
        else:
            # 创造工作表并保存
            self.workbook = pxl.Workbook()
            self.workbook.save(filePath)
        self.workbook = pxl.load_workbook(self.filepath,
                                          read_only=self.readWrite)
        return self.workbook

    def createWorkbook(self, filePath: str):
        """创建工作表

        Args:
            filePath (str): 文件路径
        """
        pxl.Workbook().save(filePath)

    def OpenSheet(self, sheet: str = None):
        """打开工作表,默认活跃表

        Args:
            sheet (str): 工作表名 eg:sheet1
        """
        if sheet is None:
            self.sheet = self.workbook.active
        else:
            if sheet in self.workbook.sheetnames:
                self.sheet = self.workbook[sheet]
            else:
                return ValueError("输入错误")

    def GetWorkbook(self):
        """返回工作表对象
        """
        return self.workbook

    def SetWorkbook(self, Workbook):
        """设置工作表对象
        """
        self.workbook = Workbook

    def GetSheek(self):
        """返回工作表对象
        """
        return self.sheet

    def SetSheet(self, sheet):
        """设置工作表对象
        """
        self.sheet = sheet

    def SetCell(self, Function, start: str, end: str = None):
        """指定函数对指定区域元素的cell对象进行操作

        Args:
            start (str):起始位置
            end (str): 结束位置
            Function (Function): 操作函数
        """
        if end is None:
            Function(self.sheet[start])
        else:
            # 获取区域范围
            maxCell = max(start, end)
            minCell = min(start, end)
            model = r"\d+"
            maxRow = int("".join(re.findall(model, maxCell)))
            minRow = int("".join(re.findall(model, minCell)))
            model = r"[a-zA-Z]+"
            maxColumn = "".join(re.findall(model, maxCell))
            maxColumn = column_index_from_string(maxColumn)
            minColumn = "".join(re.findall(model, minCell))
            minColumn = column_index_from_string(minColumn)
            # 获取单元格对象
            for row in range(minRow, maxRow + 1):
                for column in range(minColumn, maxColumn + 1):
                    # print(self.sheet.cell(row, column))
                    Function(self.sheet.cell(row, column))

    def LengthAdaptation(self, column: str = "selectAll"):
        """长度自适应，保证每个单元格不会显示#######
        
        Args:
            column (str):默认全部。列号，eg：A
        """
        # 找出一列中最长的将其设为这列的长度
        if self.cellLength:
            pass
        else:
            Excel.SaveDict(self)
        tempDict = dict()
        temp = set()
        model = r"[a-zA-Z]+"
        for i in self.cellLength.keys():
            x = re.findall(model, i)
            # x>1 i非法
            if x:
                if x[0] in tempDict.keys():
                    temp = copy.deepcopy(tempDict.get(x[0]))
                    temp.add(self.cellLength.get(i))
                    tempDict[x[0]] = temp
                else:
                    temp.add(self.cellLength.get(i))
                    tempDict[x[0]] = temp
        self.ColumnLength = copy.deepcopy(tempDict)
        if column == "selectAll":
            for i in self.ColumnLength.keys():
                self.sheet.column_dimensions[i].width = max(
                    self.ColumnLength.get(i)) + 4
        else:
            if column in self.ColumnLength.keys():
                self.sheet.column_dimensions[column].width = max(
                    self.ColumnLength.get(column)) + 4
            else:
                return ValueError("没有这一列")
        # sheet.column_dimensions['B'].width = len(str(sheet['B4'].value))

    def append(self, item):
        self.sheet.append(item)

    def Close(self):
        """
        关闭文件
        """
        self.workbook.close()

    def Save(self):
        self.workbook.save(self.filepath)

    def SaveDict(self):
        """
        将Excel内存转成内容——位置（eg:A1）对照的字典
        """
        dt = dict()
        for cell in self.sheet.rows:
            # >> cell <class 'tuple'> (<Cell '原数据'.A1>……<Cell '原数据'.F1>)
            long = len(cell)
            for i in range(long):
                temp = set()
                if isChinese(cell[i].value):
                    self.cellLength[cell[i].coordinate] = len(
                        str(cell[i].value)) * 3
                else:
                    self.cellLength[cell[i].coordinate] = len(
                        str(cell[i].value))
                # 若字典有值则添加，无则创建
                if cell[i].value in dt.keys():
                    temp = copy.deepcopy(dt.get(cell[i].value))
                    temp.add(cell[i].coordinate)
                    dt[cell[i].value] = temp
                else:
                    temp.add(cell[i].coordinate)
                    dt[cell[i].value] = temp
        self.keyValue = copy.deepcopy(dt)

    def GetRanks(self, value: str) -> set:
        """返回值所在行列号
        
        Args:
            value (str):要查询的值(execl中要有，不然空集合)
        """
        if self.keyValue:
            pass
        else:
            Excel.SaveDict(self)
        if value in self.keyValue.keys():
            return self.keyValue.get(value)

    def GetpositionValue(self, position: str) -> str:
        """返回指定位置的值
        
        Args:
            position (str): 位置 eg：A1
        """
        return self.sheet[position].value

    def GetRow(self, row: str) -> tuple:
        """返回指定行的内容

        Args:
            row (str): 行号 eg:A1 or 1

        Returns:
            tuple: [description]
        """
        # 正则提取行号，遍历列
        out = list()
        model = r"\d+"
        row = "".join(re.findall(model, row))
        column = list()
        for i in range(1, self.sheet.max_column + 1):
            column.append(get_column_letter(i) + row)
        for position in column:
            out.append(self.sheet[position].value)
        out = tuple(out)
        return out

    def GetColumn(self, column: str) -> tuple:
        """返回指定列的内容

        Args:
            column (str): 列号 eg:A1 or A

        Returns:
            tuple: [description]
        """
        out = list()
        row = list()
        model = r"[a-zA-Z]+"
        column = "".join(re.findall(model, column))
        for i in range(1, self.sheet.max_row + 1):
            row.append(column + str(i))
        for position in row:
            out.append(self.sheet[position].value)
        out = tuple(out)
        return out

    def GetDesignatedAreaValue(self, start: str, end: str) -> tuple:
        """返回指定区域的值
        
        Args:
            start (str): 开始位置 eg：A1
            end (str): 结束位置 eg：B1
        """
        temp = list()
        for i in self.sheet[start:end]:
            for cell in i:
                temp.append(cell.value)
        temp = tuple(temp)
        return temp


def isChinese(word: str) -> bool:
    """判断是不是汉字,非法输入返会None

    Args:
        word (str): 要验证的文字 eg：汉族，A，1

    Returns:
        bool: 是汉字则是True，否则是False eg：True False False
    """
    if isinstance(word, str):
        for ch in word:
            if '\u4e00' <= ch <= '\u9fff':
                return True
        return False
    else:
        return None


if __name__ == '__main__':
    # 测试
    # excel = Excel("openpyxl作业\数据透视表.xlsx", readOnly=True)
    excel = Excel("openpyxl作业\数据透视表.xlsx")
    excel.OpenSheet("原数据")
    # excel.SaveDict()
    print("=======================")
    # 测试内容与值的对应
    for i in excel.GetRanks("苏州"):
        if excel.GetpositionValue(i) != "苏州":
            print("错误位置", i)
            print("错误值", excel.GetpositionValue(i))
    print("=========================")
    print(excel.GetDesignatedAreaValue("A1", "B1"))
    print(type(excel.GetpositionValue("A1")))
    print("=======================")
    print(excel.GetRow("1"))
    print(excel.GetRow("A1"))
    print("========================")
    print(excel.GetColumn("A1"))
    print()
    print(excel.GetColumn("A"))
    excel.LengthAdaptation()